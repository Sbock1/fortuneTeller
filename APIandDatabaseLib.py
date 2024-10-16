from alpha_vantage.fundamentaldata import FundamentalData
from alpha_vantage.timeseries import TimeSeries
import matplotlib.pyplot as plt
import seaborn as sb
import pandas as pd
import numpy as np
import datetime as dt
import time
import sqlite3 as sql
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.dataframe import dataframe_to_rows

import tkinter as tk
import os

################################################################## 
# Author: Sebastian Bock (Sbock)
# Global variables based on API requirements
API_key="HNNMDBOG55BREC5P" #Account API key to validate access
time_delay = 1 #The free API access only allows 25 requests per day without any delay, but for test
FundamentalDataAPIPull = FundamentalData(API_key,output_format='pandas')
TimeSeriesPull = TimeSeries(key=API_key, output_format="pandas")
# IMPORTANT:
# The AlphaVantage API documentation can be found here:
# https://www.alphavantage.co/documentation/
##################################################################


##Update first row -> needs to be called separately, not included yet in general API request download -> TO-DO: Better implementation
def insertFirstRowColumnNames():
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_company_overview(stock)[0]
      
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists, if not create it
    if "Overview" not in book.sheetnames:
        balanceSheet = book.create_sheet("Overview")
    else:
        balanceSheet = book["Overview"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)

def insertFirstRowColumnNamesBalanceQuarterly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_balance_sheet_quarterly(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Balance_Quarterly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Balance_Quarterly")
    else:
        balanceSheet = book["Balance_Quarterly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def insertFirstRowColumnNamesBalanceYearly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_balance_sheet_annual(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Balance_Yearly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Balance_Yearly")
    else:
        balanceSheet = book["Balance_Yearly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()

def insertFirstRowColumnNamesIncomeQuarterly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_income_statement_quarterly(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Income_Quarterly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Income_Quarterly")
    else:
        balanceSheet = book["Income_Quarterly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def insertFirstRowColumnNamesIncomeYearly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_income_statement_annual(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Income_Yearly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Income_Yearly")
    else:
        balanceSheet = book["Income_Yearly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def insertFirstRowColumnNamesStockDaily():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = TimeSeriesPull.get_daily(stock, outputsize="full")[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    # Convert index to a column to preserve the datetime of the actual stock value
    firstRowData.reset_index(inplace=True)

    # Rename columns that start with numbers (e.g., '4. close' -> 'Close')
    column_rename_mapping = {
        'index': 'Date',
        '1. open': 'Open',
        '2. high': 'High',
        '3. low': 'Low',
        '4. close': 'Close',
        '5. volume': 'Volume'
    }
    firstRowData.rename(columns=column_rename_mapping, inplace=True)
    
    # Konvertieren der 'Datetime'-Spalte in 'Date', um nur das Datum zu erhalten
    firstRowData['date'] = pd.to_datetime(firstRowData['date']).dt.date   

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Stocks_Daily" not in book.sheetnames:
        balanceSheet = book.create_sheet("Stocks_Daily")
    else:
        balanceSheet = book["Stocks_Daily"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def insertFirstRowColumnNamesCashflowYearly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_cash_flow_annual(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Cashflow_Yearly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Cashflow_Yearly")
    else:
        balanceSheet = book["Cashflow_Yearly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()

def insertFirstRowColumnNamesCashflowQuarterly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_cash_flow_quarterly(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Cashflow_Quarterly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Cashflow_Quarterly")
    else:
        balanceSheet = book["Cashflow_Quarterly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def deleletExcelPredefinedSheet():
    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if predefined worksheet exists and delete it
    if "Tabelle1" in book.sheetnames:
        del book["Tabelle1"]
    elif "Sheet1" in book.sheetnames:
        del book["Sheet"]

    # Speichere die Änderungen
    book.save(excel_path)




############################# TO-DO-
def createAnalysisYearlyTable():
    
    # Connect to database
    conn = sql.connect("mainDatabase.db")
    cursor = conn.cursor()

    # Create table if NOT existing in Database
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Analysis_Yearly (
            Symbol_Date TEXT,
            Symbol TEXT,
            fiscalDateEnding DATE,
            Earnings_Per_Share REAL,
            Price_to_Earnings_Ratio REAL,
            Dividend_Yield REAL,
            Market_Cap REAL,
            Price_to_Book_Ratio REAL,
            Return_on_Capital_Employed REAL,
                   
            Gross_Profit_Margin REAL,
            Operating_Profit_Margin REAL,
            Net_Profit_Margin REAL,
            Return_on_Assets REAL,
            Return_on_Equity REAL,
            Return_on_Investment REAL,
            Equity_Ratio REAL,
            Debt_Ratio REAL,
            Interest_Coverage_Ratio REAL,
            Debt_Service_Coverage_Ratio REAL,
            Leverage_Ratio REAL,
            Debt_to_Capital_Ratio REAL,
            Short_term_Debt_Ratio REAL,
            Long_term_Debt_Ratio REAL,
            Current_Ratio REAL,
            Quick_Ratio REAL,
            Debt_to_Equity_Ratio REAL,
            Working_Capital REAL,
            Working_Capital_Ratio REAL,
            Net_Working_Capital_Ratio REAL,
            Days_Inventory_Outstanding REAL,
            Days_Sales_Outstanding REAL,
            Days_Payable_Outstanding REAL,
            Cash_Conversion_Cycle REAL,
            Cash_Ratio REAL,
            Operating_Cashflow_Ratio REAL,
            Asset_Turnover REAL,
            Equity_to_fixed_Assets_Ratio REAL,
            Extended_Coverage_Ratio REAL,
            Extended_Asset_Coverage_Ratio REAL,
            PRIMARY KEY (Symbol, fiscalDateEnding)
        );
    ''')

    # Daten einfügen
    cursor.execute('''
        INSERT OR REPLACE INTO 
            Analysis_Yearly (
                Symbol_Date,
                Symbol,
                fiscalDateEnding, 
                Earnings_Per_Share,
                Price_to_Earnings_Ratio,
                Dividend_Yield,
                Market_Cap,
                Price_to_Book_Ratio,
                Return_on_Capital_Employed,

                Gross_Profit_Margin,
                Operating_Profit_Margin,
                Net_Profit_Margin,
                Return_on_Assets,
                Return_on_Equity,
                Return_on_Investment,
                Equity_Ratio, 
                Debt_Ratio,
                Interest_Coverage_Ratio,
                Debt_Service_Coverage_Ratio,
                Leverage_Ratio,
                Debt_to_Capital_Ratio,
                Short_term_Debt_Ratio,
                Long_term_Debt_Ratio,
                Current_Ratio,
                Quick_Ratio,
                Debt_to_Equity_Ratio,  
                Working_Capital,
                Working_Capital_Ratio,
                Net_Working_Capital_Ratio,
                Days_Inventory_Outstanding,
                Days_Sales_Outstanding,
                Days_Payable_Outstanding,
                Cash_Conversion_Cycle,
                Cash_Ratio,
                Operating_Cashflow_Ratio,
                Asset_Turnover,
                Equity_to_fixed_Assets_Ratio,
                Extended_Coverage_Ratio,
                Extended_Asset_Coverage_Ratio 
            )
        SELECT
            b.Symbol || ' ' || b.fiscalDateEnding AS Symbol_Date,
            b.Symbol,
            b.fiscalDateEnding,
            ROUND((i.netIncome * 1.00) / b.commonstockSharesOutstanding, 3) AS Earnings_Per_Share,
            ROUND((sd.Close * 1.00) / ((i.netIncome * 1.00) / b.commonstockSharesOutstanding), 3) AS Price_to_Earnings_Ratio,
            ROUND(((c.dividendPayout * 1.00) / b.commonstockSharesOutstanding) / sd.Close, 4) AS Dividend_Yield,
            ROUND(((sd.Close * 1.00) * b.commonstockSharesOutstanding), 0) AS Market_Cap,
            ROUND((sd.Close * 1.00) / (b.totalShareholderEquity * 1.00 / b.commonstockSharesOutstanding), 3) AS Price_to_Book_Ratio,
            ROUND(i.operatingIncome * 1.00 / (b.totalAssets - b.totalCurrentLiabilities), 3) AS Return_on_Capital_Employed,
                      
            ROUND((i.grossProfit * 1.00) / i.totalRevenue, 3) AS Gross_Profit_Margin,
            ROUND((i.operatingIncome * 1.00) / i.totalRevenue, 3) AS Operating_Profit_Margin,
            ROUND((i.netIncome * 1.00) / i.totalRevenue, 3) AS Net_Profit_Margin,
            ROUND((i.netIncome * 1.00) / b.totalAssets, 3) AS Return_on_Assets,
            ROUND((i.netIncome * 1.00) / b.totalShareholderEquity, 3) AS Return_on_Equity,
            ROUND((i.netIncome * 1.00) / (b.totalAssets - b.totalLiabilities), 3) AS Return_on_Investment,
                    
            ROUND((b.totalShareholderEquity * 1.00) / b.totalAssets, 3) AS Equity_Ratio,
            ROUND((b.totalLiabilities * 1.00) / b.totalAssets, 3) AS Debt_Ratio,
            ROUND((i.operatingIncome * 1.00) / i.interestExpense, 3) AS Interest_Coverage_Ratio,
            ROUND((i.operatingIncome * 1.00) / i.interestAndDebtExpense, 3) AS Debt_Service_Coverage_Ratio,
            ROUND(b.totalLiabilities * 1.00 / b.totalShareholderEquity, 3) AS Leverage_Ratio,
            ROUND(b.totalLiabilities * 1.00 / (b.totalLiabilities + b.totalShareholderEquity), 3) AS Debt_to_Capital_Ratio,
            ROUND(b.totalCurrentLiabilities * 1.00 / b.totalAssets, 3) AS Short_term_Debt_Ratio,
            ROUND(b.longTermDebt * 1.00 / b.totalAssets, 3) AS Long_term_Debt_Ratio,
       
            ROUND((b.totalCurrentAssets * 1.00) / b.totalCurrentLiabilities, 3) AS Current_Ratio,
            ROUND(((b.totalCurrentAssets - b.inventory) * 1.00) / b.totalCurrentLiabilities, 3) AS Quick_Ratio,
            ROUND((b.totalLiabilities * 1.00) / b.totalShareholderEquity, 3) AS Debt_to_Equity_Ratio,
            (b.totalCurrentAssets - b.totalCurrentLiabilities) AS Working_Capital,
            ROUND((b.totalCurrentAssets * 1.00) / b.totalCurrentLiabilities, 3) AS Working_Capital_Ratio,
            ROUND(((b.totalCurrentAssets - b.totalCurrentLiabilities) * 1.00) / b.totalCurrentLiabilities, 3) AS Net_Working_Capital_Ratio,
            ROUND(((b.inventory * 1.00) / i.costofGoodsAndServicesSold) * 365, 3) AS Days_Inventory_Outstanding,
            ROUND(((b.currentNetReceivables * 1.00) / i.totalRevenue) * 365, 3) AS Days_Sales_Outstanding,
            ROUND(((b.currentAccountsPayable * 1.00) / i.costofGoodsAndServicesSold) * 365, 3) AS Days_Payable_Outstanding,
            (ROUND(((b.inventory * 1.00) / i.costofGoodsAndServicesSold) * 365, 3) + ROUND(((b.currentNetReceivables * 1.00) / i.totalRevenue) * 365, 3) - ROUND(((b.currentAccountsPayable * 1.00) / i.costofGoodsAndServicesSold) * 365, 3)) AS Cash_Conversion_Cycle,
            ROUND((b.cashAndShortTermInvestments * 1.00) / b.totalCurrentLiabilities, 3) AS Cash_Ratio,
            ROUND((c.operatingCashflow * 1.00) / b.totalCurrentLiabilities, 3) AS Operating_Cashflow_Ratio,
            ROUND((i.totalRevenue * 1.00) / b.totalAssets, 3) AS Asset_Turnover,
            ROUND((b.totalShareholderEquity * 1.00) / b.totalNonCurrentAssets, 3) AS Equity_to_fixed_Assets_Ratio,
            ROUND((b.totalShareholderEquity + b.longTermDebt) * 1.00 / b.totalNonCurrentAssets, 3) AS Extended_Coverage_Ratio,
            ROUND((b.totalShareholderEquity + b.longTermDebt) * 1.00 / b.totalAssets, 3) AS Extended_Asset_Coverage_Ratio
        FROM 
            Balance_Yearly b
        INNER JOIN 
            Income_Yearly i
        ON 
            b.Symbol = i.Symbol AND 
            b.fiscalDateEnding = i.fiscalDateEnding
        INNER JOIN
            Cashflow_Yearly c
        ON
            b.Symbol = c.Symbol AND 
            b.fiscalDateEnding = c.fiscalDateEnding
        LEFT JOIN
            Stocks_Daily sd
        ON
            b.Symbol = sd.Symbol AND 
            sd.date = (
                SELECT MIN(sd2.date)
                FROM Stocks_Daily sd2
                WHERE sd2.Symbol = b.Symbol
                AND sd2.date >= b.fiscalDateEnding
            )
            
        WHERE 
            b.totalAssets IS NOT NULL AND
            sd.close IS NOT NULL AND
            b.commonstockSharesOutstanding IS NOT NULL AND
            b.totalCurrentLiabilities IS NOT NULL AND
            i.costofGoodsAndServicesSold IS NOT NULL AND
            b.totalShareholderEquity IS NOT NULL;
    ''')

    conn.commit()
    conn.close()

    
def createAnalysisQuarterlyTable():
    conn = sql.connect("mainDatabase.db")
    cursor = conn.cursor()

    # Create table IF NOT existing already
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Analysis_Quarterly (
            Symbol_Date TEXT,
            Symbol TEXT,
            FiscalDateEnding DATE,
            Net_Income REAL,
            Operating_Cashflow REAL,
            Net_Change_in_Cash REAL,
            Free_Cashflow REAL,
            Cash_From_Financing REAL,
            Cash_From_Investing REAL,
            Earnings_Per_Share REAL,
            Price_to_Earnings_Ratio REAL,
            Dividend_Yield REAL,
            Market_Cap REAL,
            Price_to_Book_Ratio REAL,
            Return_on_Capital_Employed REAL,      
            Gross_Profit_Margin REAL,
            Operating_Profit_Margin REAL,
            Net_Profit_Margin REAL,
            Return_on_Assets REAL,
            Return_on_Equity REAL,
            Return_on_Investment REAL,
            Equity_Ratio REAL,
            Debt_Ratio REAL,
            Interest_Coverage_Ratio REAL,
            Debt_Service_Coverage_Ratio REAL,
            Leverage_Ratio REAL,
            Debt_to_Capital_Ratio REAL,
            Short_term_Debt_Ratio REAL,
            Long_term_Debt_Ratio REAL,
            Current_Ratio REAL,
            Quick_Ratio REAL,
            Debt_to_Equity_Ratio REAL,
            Working_Capital REAL,
            Working_Capital_Ratio REAL,
            Net_Working_Capital_Ratio REAL,
            Days_Inventory_Outstanding REAL,
            Days_Sales_Outstanding REAL,
            Days_Payable_Outstanding REAL,
            Cash_Conversion_Cycle REAL,
            Cash_Ratio REAL,
            Operating_Cashflow_Ratio REAL,
            Asset_Turnover REAL,
            Equity_to_fixed_Assets_Ratio REAL,
            Extended_Coverage_Ratio REAL,
            Extended_Asset_Coverage_Ratio REAL,
            PRIMARY KEY (Symbol, fiscalDateEnding)
        );
    ''')

    # Daten einfügen
    cursor.execute('''
        INSERT OR REPLACE INTO 
            Analysis_Quarterly (
                Symbol_Date,
                Symbol,
                FiscalDateEnding, 
                Net_Income,
                Operating_Cashflow,
                Net_Change_in_Cash,
                Free_Cashflow,
                Cash_From_Financing,
                Cash_From_Investing,
                Earnings_Per_Share,
                Price_to_Earnings_Ratio,
                Dividend_Yield,
                Market_Cap,
                Price_to_Book_Ratio,
                Return_on_Capital_Employed,
                Gross_Profit_Margin,
                Operating_Profit_Margin,
                Net_Profit_Margin,
                Return_on_Assets,
                Return_on_Equity,
                Return_on_Investment,
                Equity_Ratio, 
                Debt_Ratio,
                Interest_Coverage_Ratio,
                Debt_Service_Coverage_Ratio,
                Leverage_Ratio,
                Debt_to_Capital_Ratio,
                Short_term_Debt_Ratio,
                Long_term_Debt_Ratio,
                Current_Ratio,
                Quick_Ratio,
                Debt_to_Equity_Ratio,  
                Working_Capital,
                Working_Capital_Ratio,
                Net_Working_Capital_Ratio,
                Days_Inventory_Outstanding,
                Days_Sales_Outstanding,
                Days_Payable_Outstanding,
                Cash_Conversion_Cycle,
                Cash_Ratio,
                Operating_Cashflow_Ratio,
                Asset_Turnover,
                Equity_to_fixed_Assets_Ratio,
                Extended_Coverage_Ratio,
                Extended_Asset_Coverage_Ratio 
            )
        SELECT
            b.Symbol || ' ' || b.fiscalDateEnding AS Symbol_Date,
            b.Symbol,
            b.fiscalDateEnding,
            ROUND(i.netIncome, 0) AS Net_Income,
            ROUND(c.operatingCashflow, 0) AS Operating_Cashflow,
            ROUND((c.operatingCashflow + c.cashflowFromFinancing + c.cashflowFromInvestment), 0) AS Net_change_in_Cash,
            ROUND((c.operatingCashflow - c.capitalExpenditures), 0) AS Free_Cashflow,
            ROUND(c.cashflowFromFinancing, 0) AS Cash_From_Financing,
            ROUND(c.cashflowFromInvestment, 0) AS Cash_From_Investing,
            ROUND((i.netIncome * 1.00) / b.commonstockSharesOutstanding, 3) AS Earnings_Per_Share,
            ROUND((sd.Close * 1.00) / ((i.netIncome * 1.00) / b.commonstockSharesOutstanding), 3) AS Price_to_Earnings_Ratio,
            ROUND(((c.dividendPayout * 1.00) / b.commonstockSharesOutstanding) / sd.Close, 4) AS Dividend_Yield,
            ROUND(((sd.Close * 1.00) * b.commonstockSharesOutstanding), 0) AS Market_Cap,
            ROUND((sd.Close * 1.00) / (b.totalShareholderEquity * 1.00 / b.commonstockSharesOutstanding), 3) AS Price_to_Book_Ratio,
            ROUND(i.operatingIncome * 1.00 / (b.totalAssets - b.totalCurrentLiabilities), 3) AS Return_on_Capital_Employed,
                      
            ROUND((i.grossProfit * 1.00) / i.totalRevenue, 3) AS Gross_Profit_Margin,
            ROUND((i.operatingIncome * 1.00) / i.totalRevenue, 3) AS Operating_Profit_Margin,
            ROUND((i.netIncome * 1.00) / i.totalRevenue, 3) AS Net_Profit_Margin,
            ROUND((i.netIncome * 1.00) / b.totalAssets, 3) AS Return_on_Assets,
            ROUND((i.netIncome * 1.00) / b.totalShareholderEquity, 3) AS Return_on_Equity,
            ROUND((i.netIncome * 1.00) / (b.totalAssets - b.totalLiabilities), 3) AS Return_on_Investment,
                    
            ROUND((b.totalShareholderEquity * 1.00) / b.totalAssets, 3) AS Equity_Ratio,
            ROUND((b.totalLiabilities * 1.00) / b.totalAssets, 3) AS Debt_Ratio,
            ROUND((i.operatingIncome * 1.00) / i.interestExpense, 3) AS Interest_Coverage_Ratio,
            ROUND((i.operatingIncome * 1.00) / i.interestAndDebtExpense, 3) AS Debt_Service_Coverage_Ratio,
            ROUND(b.totalLiabilities * 1.00 / b.totalShareholderEquity, 3) AS Leverage_Ratio,
            ROUND(b.totalLiabilities * 1.00 / (b.totalLiabilities + b.totalShareholderEquity), 3) AS Debt_to_Capital_Ratio,
            ROUND(b.totalCurrentLiabilities * 1.00 / b.totalAssets, 3) AS Short_term_Debt_Ratio,
            ROUND(b.longTermDebt * 1.00 / b.totalAssets, 3) AS Long_term_Debt_Ratio,
       
            ROUND((b.totalCurrentAssets * 1.00) / b.totalCurrentLiabilities, 3) AS Current_Ratio,
            ROUND(((b.totalCurrentAssets - b.inventory) * 1.00) / b.totalCurrentLiabilities, 3) AS Quick_Ratio,
            ROUND((b.totalLiabilities * 1.00) / b.totalShareholderEquity, 3) AS Debt_to_Equity_Ratio,
            (b.totalCurrentAssets - b.totalCurrentLiabilities) AS Working_Capital,
            ROUND((b.totalCurrentAssets * 1.00) / b.totalCurrentLiabilities, 3) AS Working_Capital_Ratio,
            ROUND(((b.totalCurrentAssets - b.totalCurrentLiabilities) * 1.00) / b.totalCurrentLiabilities, 3) AS Net_Working_Capital_Ratio,
            ROUND(((b.inventory * 1.00) / i.costofGoodsAndServicesSold) * 365, 3) AS Days_Inventory_Outstanding,
            ROUND(((b.currentNetReceivables * 1.00) / i.totalRevenue) * 365, 3) AS Days_Sales_Outstanding,
            ROUND(((b.currentAccountsPayable * 1.00) / i.costofGoodsAndServicesSold) * 365, 3) AS Days_Payable_Outstanding,
            (ROUND(((b.inventory * 1.00) / i.costofGoodsAndServicesSold) * 365, 3) + ROUND(((b.currentNetReceivables * 1.00) / i.totalRevenue) * 365, 3) - ROUND(((b.currentAccountsPayable * 1.00) / i.costofGoodsAndServicesSold) * 365, 3)) AS Cash_Conversion_Cycle,
            ROUND((b.cashAndShortTermInvestments * 1.00) / b.totalCurrentLiabilities, 3) AS Cash_Ratio,
            ROUND((c.operatingCashflow * 1.00) / b.totalCurrentLiabilities, 3) AS Operating_Cashflow_Ratio,
            ROUND((i.totalRevenue * 1.00) / b.totalAssets, 3) AS Asset_Turnover,
            ROUND((b.totalShareholderEquity * 1.00) / b.totalNonCurrentAssets, 3) AS Equity_to_fixed_Assets_Ratio,
            ROUND((b.totalShareholderEquity + b.longTermDebt) * 1.00 / b.totalNonCurrentAssets, 3) AS Extended_Coverage_Ratio,
            ROUND((b.totalShareholderEquity + b.longTermDebt) * 1.00 / b.totalAssets, 3) AS Extended_Asset_Coverage_Ratio
        FROM 
            Balance_Quarterly b
        INNER JOIN 
            Income_Quarterly i
        ON 
            b.Symbol = i.Symbol AND 
            b.fiscalDateEnding = i.fiscalDateEnding
        INNER JOIN
            Cashflow_Quarterly c
        ON
            b.Symbol = c.Symbol AND 
            b.fiscalDateEnding = c.fiscalDateEnding
        LEFT JOIN
            Stocks_Daily sd
        ON
            b.Symbol = sd.Symbol AND 
            sd.date = (
                SELECT MIN(sd2.date)
                FROM Stocks_Daily sd2
                WHERE sd2.Symbol = b.Symbol
                AND sd2.date >= b.fiscalDateEnding
            )
            
        WHERE 
            b.totalAssets IS NOT NULL AND
            sd.close IS NOT NULL AND
            b.commonstockSharesOutstanding IS NOT NULL AND
            b.totalCurrentLiabilities IS NOT NULL AND
            i.costofGoodsAndServicesSold IS NOT NULL AND
            b.totalShareholderEquity IS NOT NULL;
    ''')

    conn.commit()
    conn.close()


def cleaningDatabaseNulltoZero():
    conn = sql.connect("mainDatabase.db")
    cursor = conn.cursor()

    # Request to fetch all columns in the table
    cursor.execute('''
    SELECT name FROM sqlite_master WHERE type='table' AND name != "Overview";
    ''')
    tables = [row[0] for row in cursor.fetchall()]

    for table in tables:
        cursor.execute(f'''
            PRAGMA table_info({table});
        ''')
        columns = cursor.fetchall()

        # Dynamic SQL generation to replace NULL with zero
        for column in columns:
            column_name = column[1]
            cursor.execute(f"UPDATE {table} SET {column_name} = 0 WHERE {column_name} IS NULL OR {column_name} = 'None';")
        
    conn.commit()
    conn.close()


def addingPrimaryKeyColumn():
    conn = sql.connect("mainDatabase.db")
    cursor = conn.cursor()

    # Request to fetch all columns in the table
    cursor.execute('''
    SELECT name FROM sqlite_master WHERE type='table' AND name != "Overview" AND name != "Analysis_Yearly" AND name != "Analysis_Quarterly";
    ''')
    tables = [row[0] for row in cursor.fetchall()]
    print(tables)

    for table in tables:
        cursor.execute(f"PRAGMA table_info({table})")
        columns = cursor.fetchall()

        print(table)
        print(columns[1][1])

        if "Symbol_Date" not in [column[1] for column in columns]:
            cursor.execute(f'''
                ALTER TABLE {table} ADD COLUMN Symbol_Date TEXT;
            ''')
        if "fiscalDateEnding" in [column[1] for column in columns]:
            cursor.execute(f'''
                UPDATE {table} SET Symbol_Date = Symbol || ' ' || fiscalDateEnding;
            ''')
        if "date" in [column[1] for column in columns]:
            cursor.execute(f'''
                UPDATE {table} SET Symbol_Date = Symbol || ' ' || date;
            ''')

    conn.commit()
    conn.close()

#############################


def writeToDataBase(mainExcel, database):
    conn = sql.connect("mainDatabase.db")
    mainExcel.to_sql(database, conn, if_exists="append", index=False)
    

    conn.close()




# Reads any column of an excel worksheet and gives out a list of the contents of the rows of that worksheet
def loadOneColumnRowDataAsList(filename, sheetname, column):
    listFromColumn = pd.read_excel(filename, sheet_name=sheetname, usecols=column, header=None, index_col=0) #Read column A / No header - Take first row NOT as replacement header / No index / returns Dataframe
    listFromColumn = listFromColumn.index.tolist() # Convert Dataframe to list in order for easier processing later / use index as the column A is seen as index column
    return listFromColumn


# Compare mainList and deleteList, delete the items in mainList, which are not in deleteList, and return modified mainList
# Compare mainList and addList, add the items to mainList, which are in addList, and return modified mainList
# The goal of the function is helping with having a uptodate SymbolList
def compareDeleteAddListWithMainList(mainList, deleteList=None, addList=None):
    if deleteList == None:
        deleteList = []
    if addList == None:
        addList =[]
    
    # KEEP IN MIND: First the deletion is done, THEN the addition!
    mainList = [item for item in mainList if item not in deleteList]
    mainList.extend([item for item in addList if item not in mainList])
    mainList.sort()

    return mainList


def getExcelSheetInformation(filename, sheetname):
    mainSheetDataframe = pd.read_excel(filename, sheetname)


    excelSymbolsExisting = list(set(mainSheetDataframe["Symbol"]))
    excelQuartersExisting = list(mainSheetDataframe["Downloaded_at_quarter"])
    
    return excelSymbolsExisting, excelQuartersExisting



def checkSymbolCurrentQuarterExisting(symbolList, excelSymbolsExisting, excelQuartersExisting):
    # This func checks three things:
    # 1. Which symbols (stocks) are already in the main excel written
    # 2. Which symbols are already in the main excel existing, but the download quarter is old, and they needRefresh
    # 3. Which symbols do not exist anymore, meaning which companies are not longer on the stock exchange registered
    # This func does not eliminate any stock from the main stock list, but it gives the information to Writer functions
    existingSymbols = []
    symbolsNeedRefresh = []
    updateNotExistingSymbols = []
    
    if excelSymbolsExisting != []:
        for symbol in symbolList:
            if symbol in excelSymbolsExisting:
                rowOnWorksheet = excelSymbolsExisting.index(symbol)
                if timestampQuarter() == excelQuartersExisting[rowOnWorksheet]:
                    existingSymbols.append(symbol + " in" + str(timestampQuarter()) +" uptodate")
                else:
                    symbolsNeedRefresh.append([symbol, rowOnWorksheet])
            else:
                updateNotExistingSymbols.append(symbol)
                
        return existingSymbols, symbolsNeedRefresh, updateNotExistingSymbols

    else:
        updateNotExistingSymbols=symbolList

        return existingSymbols, symbolsNeedRefresh, updateNotExistingSymbols



def APIRequestDelay():
    time.sleep(time_delay)

    return None


def timestampToday():
    # Get the current datetime
    today = pd.to_datetime(dt.datetime.now())

    return today


def timestampQuarter():
    # Get the current quarter
    today = pd.to_datetime(dt.datetime.today())
    currentQuarter = (today.month - 1) // 3 + 1

    return currentQuarter


def identifyLastColumnWithContents(DataFrame):
    # To identify the last column, the DataFrame is given as input, than it is check which cells hold at not a NaN value
    # than the columns with at least one not NaN value are marked true. The last one of these is the output translated into a number,
    # this number is returned by this function.
    lastColumnWithContentsName = DataFrame.notna().any().index[-1]
    lastColumnWithContentsNumber = DataFrame.columns.get_loc(lastColumnWithContentsName)

    return lastColumnWithContentsNumber


def updateCompanyOverview(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockOverviewData = pd.DataFrame()
    counter = 0

    for elem in updateNotExistingSymbols:
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockOverviewData = FundamentalDataAPIPull.get_company_overview(elem)[0] 
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockOverviewData["Downloaded_at_datetime"] = timestampToday()
            stockOverviewData["Downloaded_at_quarter"] = timestampQuarter()

            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockOverviewData, pd.DataFrame):
                stockOverviewData = pd.DataFrame(stockOverviewData)

            allStockOverviewData = pd.concat([allStockOverviewData, stockOverviewData], ignore_index=True)

            APIRequestDelay()
        except ValueError:
            stocksNotExisting.append(elem)
            print("Error" + str(counter) + " " + str(elem))
            counter = counter + 1
            APIRequestDelay()
            pass
    
    return allStockOverviewData, stocksNotExisting

def update_balance_sheet_quarterly(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockBalanceQuarterlyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_balance_sheet_quarterly(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            stockBalanceData.insert(0, "Symbol", stock)

            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            allStockBalanceQuarterlyData = pd.concat([allStockBalanceQuarterlyData, stockBalanceData], ignore_index=True)
            
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            APIRequestDelay()
            pass
        
    return allStockBalanceQuarterlyData , stocksNotExisting 


def update_balance_sheet_annual(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockBalanceAnnuallyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_balance_sheet_annual(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            stockBalanceData.insert(0, "Symbol", stock)

            allStockBalanceAnnuallyData = pd.concat([allStockBalanceAnnuallyData, stockBalanceData], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        
    return allStockBalanceAnnuallyData, stocksNotExisting 


def update_income_statement_quarterly(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockIncomeQuarterlyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_income_statement_quarterly(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            stockBalanceData.insert(0, "Symbol", stock)

            allStockIncomeQuarterlyData = pd.concat([allStockIncomeQuarterlyData, stockBalanceData], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        
    return allStockIncomeQuarterlyData, stocksNotExisting  



def update_income_statement_annual(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockIncomeAnnuallyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_income_statement_annual(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            stockBalanceData.insert(0, "Symbol", stock)

            allStockIncomeAnnuallyData = pd.concat([allStockIncomeAnnuallyData, stockBalanceData], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        
    return allStockIncomeAnnuallyData, stocksNotExisting  


def updateCashflowStatementAnually(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockCashflowAnuallyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_cash_flow_annual(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            stockBalanceData.insert(0, "Symbol", stock)

            allStockCashflowAnuallyData = pd.concat([allStockCashflowAnuallyData, stockBalanceData], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        
    return allStockCashflowAnuallyData, stocksNotExisting 
    

def updateCashflowStatementQuarterly(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockCashflowAnuallyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_cash_flow_quarterly(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            stockBalanceData.insert(0, "Symbol", stock)

            allStockCashflowAnuallyData = pd.concat([allStockCashflowAnuallyData, stockBalanceData], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        
    return allStockCashflowAnuallyData, stocksNotExisting 

def getTimeSeriesData(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockTimeSeries = pd.DataFrame()
    counter = 0

    for stock in updateNotExistingSymbols:
        try:
            # Abruf der täglichen Zeitreihe
            dailySeriesPerStock = TimeSeriesPull.get_daily(stock, outputsize="full")[0]

            # Sicherstellen, dass es sich um ein DataFrame handelt
            if not isinstance(dailySeriesPerStock, pd.DataFrame):
                dailySeriesPerStock = pd.DataFrame(dailySeriesPerStock)

            # Datum vom Index in eine Spalte umwandeln
            dailySeriesPerStock.reset_index(inplace=True)

            # Hinzufügen von Zeitstempeln für den Download
            dailySeriesPerStock["Downloaded_at_datetime"] = timestampToday()
            dailySeriesPerStock["Downloaded_at_quarter"] = timestampQuarter()
                   
            # Umbenennen der Spalten, die mit Zahlen beginnen, und 'Datetime' in 'Date' umwandeln
            column_rename_mapping = {
                'index': 'Date',
                '1. open': 'Open',
                '2. high': 'High',
                '3. low': 'Low',
                '4. close': 'Close',
                '5. volume': 'Volume'
            }
            dailySeriesPerStock.rename(columns=column_rename_mapping, inplace=True)

            # Umwandlung von 'Date' in reines Datum (ohne Zeitstempel)
            dailySeriesPerStock["date"] = pd.to_datetime(dailySeriesPerStock["date"]).dt.date

            # Einfügen des Symbols in die erste Spalte
            dailySeriesPerStock.insert(0, "Symbol", stock)

            # Zusammenführen mit dem allStockTimeSeries DataFrame
            allStockTimeSeries = pd.concat([allStockTimeSeries, dailySeriesPerStock], ignore_index=True)

            # API-Abfrage-Verzögerung
            APIRequestDelay()

        except ValueError:
            stocksNotExisting.append(stock)
            print(f"Error {counter}: {stock}")
            counter += 1
            pass

    return allStockTimeSeries, stocksNotExisting  


""" TO-DO: Check existing data and expand it
def refresh(excel_data, refresh_list):
    today = pd.to_datetime(dt.datetime.today())
    today_quarter = int(today.quarter)
    for elem in refresh_list:
        try:
            stock1 = fd.get_company_overview(elem)[0]
            stock1.insert(59, "Downloaded_at", pd.to_datetime(dt.datetime.today()))
            stock1.insert(60, "Downloaded_quarter", today_quarter)
            excel_data.loc[excel_data["Symbol"] == elem] = stock_transform(stock1)
            time.sleep(time_delay)
        except ValueError:
            excel_data_refreshed = excel_data
            return excel_data_ref
    excel_data_refreshed = excel_data
    return excel_data_ref

def load_stock_price_yf(stock_list_new):
    stock_price_df = yf.download(stock_list_new, period="1d")
    stock_price_df_t = stock_price_df["Close"].T
    stock_price_df_t.insert(1, "Date", pd.to_datetime(dt.datetime.today()))
    stock_price_df_t1 = stock_price_df_t.reset_index()
    stock_price_df_t1.columns = ["Symbol","Price","Date"]
    
    return stock_price_df_t1
"""
           
def writeToExcel(mainExcel, worksheet):
    
    excelPath = "output.xlsx"

    book = load_workbook(excelPath)

    balanceSheet = book[worksheet]

    for row in dataframe_to_rows(mainExcel, index=False, header=False):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excelPath)


def readFromDataBase(database):
    # Connection to Database
    conn = sql.connect("mainDatabase.db")

    try:
        mainSheetDataframe = pd.read_sql(f"SELECT Symbol, Downloaded_at_quarter FROM {database}", conn)
        
        excelSymbolsExisting = list(set(mainSheetDataframe["Symbol"]))
        excelQuartersExisting = list(mainSheetDataframe["Downloaded_at_quarter"])
        
        return excelSymbolsExisting, excelQuartersExisting

    except:

        excelSymbolsExisting = ["A"]
        excelQuartersExisting = timestampQuarter()

        return excelSymbolsExisting, excelQuartersExisting

    finally:
        conn.close()


def writeToDataBase(mainExcel, database):

    conn = sql.connect("mainDatabase.db")
    mainExcel.to_sql(database, conn, if_exists="append", index=False)
    

    conn.close()


def sortDatabaseBySymbolName(table):
    
    conn = sql.connect("mainDatabase.db")

    cursor = conn.cursor()

    # Create temp table differently sorted
    cursor.execute(f"CREATE TABLE sorted_{table} AS SELECT * FROM {table} ORDER BY Symbol ASC")
    
    # Delete old table
    cursor.execute(f"DROP TABLE {table}")
    
    cursor.execute(f"ALTER TABLE sorted_{table} RENAME TO {table}")


    conn.commit()

    conn.close()



# NOT USED: Implementation as is, leads to deletion of Symbol entries in Stock_symbol_list, because the API request does get canceled from provide due to exhaustion of free API requests per day.
# TO-DO: When this particular error comes back from server, it is needed to ignore this delete func.
def deleteNoneUpdatableSymbols(symbol_list, stocksNotExisting):
    filename = "Stock_symbols_list.xlsx"
    sheetname = "Overview"
     
    # Compare symbol list against stocks not existing anymore at the stock exchange list
    filtered_symbols = [symbol for symbol in symbol_list if symbol not in stocksNotExisting]

    excelDataFrame = pd.DataFrame(filtered_symbols, columns=[0])

    excelDataFrame.to_excel(filename, sheet_name=sheetname, index=False, header=False)


