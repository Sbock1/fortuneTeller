from alpha_vantage.fundamentaldata import FundamentalData
from alpha_vantage.timeseries import TimeSeries
import matplotlib.pyplot as plt
import seaborn as sb
import pandas as pd
import numpy as np
import datetime as dt
import time
import sqlite3 as sql
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.dataframe import dataframe_to_rows

import tkinter as tk
import os

################################################################## 
# Author: Sebastian Bock (Sbock)
# Global variables based on API requirements
API_key="HNNMDBOG55BREC5P" #Account API key to validate access
time_delay = 1 #The free API access only allows 25 requests per day without any delay, but for test
FundamentalDataAPIPull = FundamentalData(API_key,output_format='pandas')
TimeSeriesPull = TimeSeries(key=API_key, output_format="pandas")
# IMPORTANT:
# The AlphaVantage API documentation can be found here:
# https://www.alphavantage.co/documentation/
##################################################################


##Update first row -> needs to be called separately, not included yet in general API request download -> TO-DO: Better implementation
def insertFirstRowColumnNames():
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_company_overview(stock)[0]
      
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists, if not create it
    if "Overview" not in book.sheetnames:
        balanceSheet = book.create_sheet("Overview")
    else:
        balanceSheet = book["Overview"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)

def insertFirstRowColumnNamesBalanceQuarterly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_balance_sheet_quarterly(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Balance_Quarterly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Balance_Quarterly")
    else:
        balanceSheet = book["Balance_Quarterly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def insertFirstRowColumnNamesBalanceYearly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_balance_sheet_annual(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Balance_Yearly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Balance_Yearly")
    else:
        balanceSheet = book["Balance_Yearly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()

def insertFirstRowColumnNamesIncomeQuarterly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_income_statement_quarterly(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Income_Quarterly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Income_Quarterly")
    else:
        balanceSheet = book["Income_Quarterly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def insertFirstRowColumnNamesIncomeYearly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_income_statement_annual(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Income_Yearly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Income_Yearly")
    else:
        balanceSheet = book["Income_Yearly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def insertFirstRowColumnNamesStockDaily():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = TimeSeriesPull.get_daily(stock, outputsize="full")[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    # Convert index to a column to preserve the datetime of the actual stock value
    firstRowData.reset_index(inplace=True)

    # Rename the column to Date
    firstRowData.rename(columns={'index':'Date'}, inplace=True)
    firstRowData["date"] = firstRowData["date"].dt.date

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Stocks_Daily" not in book.sheetnames:
        balanceSheet = book.create_sheet("Stocks_Daily")
    else:
        balanceSheet = book["Stocks_Daily"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def insertFirstRowColumnNamesCashflowYearly():
    # Load balance data from Alpha Vantage
    stock = "A"
    firstRowData = FundamentalDataAPIPull.get_cash_flow_annual(stock)[0]
    
    # Adding of additional columns at the end
    firstRowData["Downloaded_at_datetime"] = timestampToday()
    firstRowData["Downloaded_at_quarter"] = timestampQuarter()

    # Check and if not Dataframe datatype, convert to Dataframe
    if not isinstance(firstRowData, pd.DataFrame):
        firstRowData = pd.DataFrame([firstRowData])

    firstRowData.insert(0, "Symbol", stock)

    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if worksheet exists
    if "Cashflow_Yearly" not in book.sheetnames:
        balanceSheet = book.create_sheet("Cashflow_Yearly")
    else:
        balanceSheet = book["Cashflow_Yearly"]

    for row in dataframe_to_rows(firstRowData, index=False, header=True):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excel_path)
    APIRequestDelay()


def deleletExcelPredefinedSheet():
    # Path of file
    excel_path = "output.xlsx"

    # Check if file exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    book = load_workbook(excel_path)

    # Check if predefined worksheet exists and delete it
    if "Tabelle1" in book.sheetnames:
        del book["Tabelle1"]
    elif "Sheet1" in book.sheetnames:
        del book["Sheet"]

    # Speichere die Änderungen
    book.save(excel_path)


############################# TO-DO-
def createAnalysisYearlyTable():
    try:
        conn = sql.connect("mainDatabase.db")
        cursor = conn.cursor()

        # Tabelle erstellen
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS Analysis_Yearly (
                Symbol TEXT,
                fiscalDateEnding DATE,
                Gross_Profit_Margin,
                Operating_Profit_Margin,
                Net_Profit_Margin,
                Return_on_Assets,
                Return_on_Equity,
                Return_on_Investment,
                Equity_Ratio REAL,
                Debt_Ratio REAL,
                Interest_Coverage_Ratio REAL,
                Debt_Service_Coverage_Ratio REAL,
                Leverage_Ratio REAL,
                Debt_to_Capital_Ratio REAL,
                Current_Ratio REAL,
                Quick_Ratio REAL,
                Debt_to_Equity_Ratio REAL,
                Working_Capital REAL,
                Working_Capital_Ratio REAL,
                Net_Working_Capital_Ratio REAL,
                Days_Inventory_Outstanding REAL,
                Days_Sales_Outstanding REAL,
                Days_Payable_Outstanding REAL,
                Cash_Conversion_Cycle REAL,
                Cash_Ratio REAL,
                Operating_Cashflow_Ratio REAL,
                Asset_Turnover REAL,
                Equity_to_fixed_Assets_Ratio REAL,
                Extended_Coverage_Ratio REAL,
                Extended_Asset_Coverage_Ratio REAL,
                PRIMARY KEY (Symbol, fiscalDateEnding)
            );
        ''')

        # Daten einfügen
        cursor.execute('''
            INSERT OR REPLACE INTO 
                Analysis_Yearly (
                    Symbol,
                    fiscalDateEnding, 
                    Gross_Profit_Margin,
                    Operating_Profit_Margin,
                    Net_Profit_Margin,
                    Return_on_Assets,
                    Return_on_Equity,
                    Return_on_Investment,
                    Equity_Ratio, 
                    Debt_Ratio,
                    Interest_Coverage_Ratio,
                    Debt_Service_Coverage_Ratio,
                    Leverage_Ratio,
                    Debt_to_Capital_Ratio,
                    Current_Ratio,
                    Quick_Ratio,
                    Debt_to_Equity_Ratio,  
                    Working_Capital,
                    Working_Capital_Ratio,
                    Net_Working_Capital_Ratio,
                    Days_Inventory_Outstanding,
                    Days_Sales_Outstanding,
                    Days_Payable_Outstanding,
                    Cash_Conversion_Cycle,
                    Cash_Ratio,
                    Operating_Cashflow_Ratio,
                    Asset_Turnover,
                    Equity_to_fixed_Assets_Ratio,
                    Extended_Coverage_Ratio,
                    Extended_Asset_Coverage_Ratio 
                )
            SELECT
                b.Symbol,
                b.fiscalDateEnding,
                ROUND((i.grossProfit * 1.00) / i.totalRevenue, 3) AS Gross_Profit_Margin,
                ROUND((i.operatingIncome * 1.00) / i.totalRevenue, 3) AS Operating_Profit_Margin,
                ROUND((i.netIncome * 1.00) / i.totalRevenue, 3) AS Net_Profit_Margin,
                ROUND((i.netIncome * 1.00) / b.totalAssets, 3) AS Return_on_Assets,
                ROUND((i.netIncome * 1.00) / b.totalShareholderEquity, 3) AS Return_on_Equity,
                ROUND((i.netIncome * 1.00) / (b.totalAssets - b.totalLiabilities), 3) AS Return_on_Investment,
                       
                ROUND((b.totalShareholderEquity * 1.00) / b.totalAssets, 3) AS Equity_Ratio,
                ROUND((b.totalLiabilities * 1.00) / b.totalAssets, 3) AS Debt_Ratio,
                ROUND((i.operatingIncome * 1.00) / i.interestExpense, 3) AS Interest_Coverage_Ratio,
                ROUND((i.operatingIncome * 1.00) / i.interestAndDebtExpense, 3) AS Debt_Service_Coverage_Ratio,
                ROUND(b.totalLiabilities * 1.00 / b.totalShareholderEquity, 3) AS Leverage_Ratio,
                ROUND(b.totalLiabilities * 1.00 / (b.totalLiabilities + b.totalShareholderEquity), 3) AS Debt_to_Capital_Ratio,
                ROUND((b.totalCurrentAssets * 1.00) / b.totalCurrentLiabilities, 3) AS Current_Ratio,
                ROUND(((b.totalCurrentAssets - b.inventory) * 1.00) / b.totalCurrentLiabilities, 3) AS Quick_Ratio,
                ROUND((b.totalLiabilities * 1.00) / b.totalShareholderEquity, 3) AS Debt_to_Equity_Ratio,
                (b.totalCurrentAssets - b.totalCurrentLiabilities) AS Working_Capital,
                ROUND((b.totalCurrentAssets * 1.00) / b.totalCurrentLiabilities, 3) AS Working_Capital_Ratio,
                ROUND(((b.totalCurrentAssets - b.totalCurrentLiabilities) * 1.00) / b.totalCurrentLiabilities, 3) AS Net_Working_Capital_Ratio,
                ROUND(((b.inventory * 1.00) / i.costofGoodsAndServicesSold) * 365, 3) AS Days_Inventory_Outstanding,
                ROUND(((b.currentNetReceivables * 1.00) / i.totalRevenue) * 365, 3) AS Days_Sales_Outstanding,
                ROUND(((b.currentAccountsPayable * 1.00) / i.costofGoodsAndServicesSold) * 365, 3) AS Days_Payable_Outstanding,
                (ROUND(((b.inventory * 1.00) / i.costofGoodsAndServicesSold) * 365, 3) + ROUND(((b.currentNetReceivables * 1.00) / i.totalRevenue) * 365, 3) - ROUND(((b.currentAccountsPayable * 1.00) / i.costofGoodsAndServicesSold) * 365, 3)) AS Cash_Conversion_Cycle,
                ROUND((b.cashAndShortTermInvestments * 1.00) / b.totalCurrentLiabilities, 3) AS Cash_Ratio,
                ROUND((c.operatingCashflow * 1.00) / b.totalCurrentLiabilities, 3) AS Operating_Cashflow_Ratio,
                ROUND((i.totalRevenue * 1.00) / b.totalAssets, 3) AS Asset_Turnover,
                ROUND((b.totalShareholderEquity * 1.00) / b.totalNonCurrentAssets, 3) AS Equity_to_fixed_Assets_Ratio,
                ROUND((b.totalShareholderEquity + b.longTermDebt) * 1.00 / b.totalNonCurrentAssets, 3) AS Extended_Coverage_Ratio,
                ROUND((b.totalShareholderEquity + b.longTermDebt) * 1.00 / b.totalAssets, 3) AS Extended_Asset_Coverage_Ratio
            FROM 
                Balance_Yearly b
            INNER JOIN 
                Income_Yearly i
            ON 
                b.Symbol = i.Symbol AND b.fiscalDateEnding = i.fiscalDateEnding
            INNER JOIN
                Cashflow_Yearly c
            ON
                b.Symbol = c.Symbol AND b.fiscalDateEnding = c.fiscalDateEnding 
            WHERE 
                b.totalAssets IS NOT NULL AND 
                b.totalCurrentLiabilities IS NOT NULL AND
                i.costofGoodsAndServicesSold IS NOT NULL AND
                b.totalShareholderEquity IS NOT NULL;
        ''')

        conn.commit()
        print("Analysis_Yearly Tabelle erfolgreich erstellt und Daten eingefügt.")
    except Exception as e:
        print(f"Fehler bei der Erstellung der Tabelle: {e}")
    finally:
        conn.close()

    
def createAnalysisQuarterlyTable():
    conn = sql.connect("mainDatabase.db")
    cursor = conn.cursor()

    # Create table IF NOT existing already
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Analysis_Quarterly (
            Symbol TEXT,
            fiscalDateEnding DATE,
            Equity_Ratio REAL,
            Debt_Ratio REAL,
            Current_Ratio REAL,
            Quick_Ratio REAL,
            Debt_to_Equity_Ratio REAL,
            Working_Capital REAL,
            Net_Working_Capital_Ratio REAL,
            PRIMARY KEY (Symbol, fiscalDateEnding)
        );
    ''')
    # Data calculation from Balance_Yearly
    cursor.execute('''
        INSERT OR REPLACE INTO 
            Analysis_Quarterly (
                Symbol,
                fiscalDateEnding, 
                Equity_Ratio, 
                Debt_Ratio, 
                Current_Ratio, 
                Quick_Ratio,
                Debt_to_Equity_Ratio,   
                Working_Capital,
                Net_Working_Capital_Ratio
                )
        SELECT
            Symbol,
            fiscalDateEnding,
            ROUND((totalShareholderEquity * 1.00) / totalAssets, 3) AS Equity_Ratio,
            ROUND((totalLiabilities * 1.00) / totalAssets, 3) AS Debt_Ratio,
            ROUND((totalCurrentAssets * 1.00) / totalCurrentLiabilities, 3) AS Current_Ratio,
            ROUND(((totalCurrentAssets - inventory) * 1.00) / totalCurrentLiabilities, 3) AS Quick_Ratio,
            ROUND((totalLiabilities * 1.00) / totalShareholderEquity, 3) AS Debt_to_Equity_Ratio,
            (totalCurrentAssets - totalCurrentLiabilities) AS Working_Capital,
            ROUND(((totalCurrentAssets - totalCurrentLiabilities) * 1.00) / totalAssets, 3) AS Net_Working_Capital_Ratio
        FROM Balance_Yearly
        WHERE 
            totalAssets IS NOT NULL AND 
            totalCurrentLiabilities IS NOT NULL AND
            totalShareholderEquity IS NOT NULL;
    ''')
    conn.commit()
    conn.close()


#############################


def writeToDataBase(mainExcel, database):
    conn = sql.connect("mainDatabase.db")
    mainExcel.to_sql(database, conn, if_exists="append", index=False)
    

    conn.close()




# Reads any column of an excel worksheet and gives out a list of the contents of the rows of that worksheet
def loadOneColumnRowDataAsList(filename, sheetname, column):
    listFromColumn = pd.read_excel(filename, sheet_name=sheetname, usecols=column, header=None, index_col=0) #Read column A / No header - Take first row NOT as replacement header / No index / returns Dataframe
    listFromColumn = listFromColumn.index.tolist() # Convert Dataframe to list in order for easier processing later / use index as the column A is seen as index column
    return listFromColumn


# Compare mainList and deleteList, delete the items in mainList, which are not in deleteList, and return modified mainList
# Compare mainList and addList, add the items to mainList, which are in addList, and return modified mainList
# The goal of the function is helping with having a uptodate SymbolList
def compareDeleteAddListWithMainList(mainList, deleteList=None, addList=None):
    if deleteList == None:
        deleteList = []
    if addList == None:
        addList =[]
    
    # KEEP IN MIND: First the deletion is done, THEN the addition!
    mainList = [item for item in mainList if item not in deleteList]
    mainList.extend([item for item in addList if item not in mainList])
    mainList.sort()

    return mainList


def getExcelSheetInformation(filename, sheetname):
    mainSheetDataframe = pd.read_excel(filename, sheetname)


    excelSymbolsExisting = list(set(mainSheetDataframe["Symbol"]))
    excelQuartersExisting = list(mainSheetDataframe["Downloaded_at_quarter"])
    
    return excelSymbolsExisting, excelQuartersExisting



def checkSymbolCurrentQuarterExisting(symbolList, excelSymbolsExisting, excelQuartersExisting):
    # This func checks three things:
    # 1. Which symbols (stocks) are already in the main excel written
    # 2. Which symbols are already in the main excel existing, but the download quarter is old, and they needRefresh
    # 3. Which symbols do not exist anymore, meaning which companies are not longer on the stock exchange registered
    # This func does not eliminate any stock from the main stock list, but it gives the information to Writer functions
    existingSymbols = []
    symbolsNeedRefresh = []
    updateNotExistingSymbols = []
    
    if excelSymbolsExisting != []:
        for symbol in symbolList:
            if symbol in excelSymbolsExisting:
                rowOnWorksheet = excelSymbolsExisting.index(symbol)
                if timestampQuarter() == excelQuartersExisting[rowOnWorksheet]:
                    existingSymbols.append(symbol + " in" + str(timestampQuarter()) +" uptodate")
                else:
                    symbolsNeedRefresh.append([symbol, rowOnWorksheet])
            else:
                updateNotExistingSymbols.append(symbol)
                
        return existingSymbols, symbolsNeedRefresh, updateNotExistingSymbols

    else:
        updateNotExistingSymbols=symbolList

        return existingSymbols, symbolsNeedRefresh, updateNotExistingSymbols



def APIRequestDelay():
    time.sleep(time_delay)

    return None


def timestampToday():
    # Get the current datetime
    today = pd.to_datetime(dt.datetime.now())

    return today


def timestampQuarter():
    # Get the current quarter
    today = pd.to_datetime(dt.datetime.today())
    currentQuarter = (today.month - 1) // 3 + 1

    return currentQuarter


def identifyLastColumnWithContents(DataFrame):
    # To identify the last column, the DataFrame is given as input, than it is check which cells hold at not a NaN value
    # than the columns with at least one not NaN value are marked true. The last one of these is the output translated into a number,
    # this number is returned by this function.
    lastColumnWithContentsName = DataFrame.notna().any().index[-1]
    lastColumnWithContentsNumber = DataFrame.columns.get_loc(lastColumnWithContentsName)

    return lastColumnWithContentsNumber


def updateCompanyOverview(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockOverviewData = pd.DataFrame()
    counter = 0

    for elem in updateNotExistingSymbols:
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockOverviewData = FundamentalDataAPIPull.get_company_overview(elem)[0] 
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockOverviewData["Downloaded_at_datetime"] = timestampToday()
            stockOverviewData["Downloaded_at_quarter"] = timestampQuarter()

            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockOverviewData, pd.DataFrame):
                stockOverviewData = pd.DataFrame(stockOverviewData)

            allStockOverviewData = pd.concat([allStockOverviewData, stockOverviewData], ignore_index=True)

            APIRequestDelay()
        except ValueError:
            stocksNotExisting.append(elem)
            print("Error" + str(counter) + " " + str(elem))
            counter = counter + 1
            APIRequestDelay()
            pass
    
    return allStockOverviewData, stocksNotExisting

def update_balance_sheet_quarterly(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockBalanceQuarterlyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_balance_sheet_quarterly(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            stockBalanceData.insert(0, "Symbol", stock)

            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            allStockBalanceQuarterlyData = pd.concat([allStockBalanceQuarterlyData, stockBalanceData], ignore_index=True)
            
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            APIRequestDelay()
            pass
        
    return allStockBalanceQuarterlyData , stocksNotExisting 


def update_balance_sheet_annual(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockBalanceAnnuallyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_balance_sheet_annual(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            stockBalanceData.insert(0, "Symbol", stock)

            allStockBalanceAnnuallyData = pd.concat([allStockBalanceAnnuallyData, stockBalanceData], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        
    return allStockBalanceAnnuallyData, stocksNotExisting 


def update_income_statement_quarterly(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockIncomeQuarterlyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_income_statement_quarterly(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            stockBalanceData.insert(0, "Symbol", stock)

            allStockIncomeQuarterlyData = pd.concat([allStockIncomeQuarterlyData, stockBalanceData], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        
    return allStockIncomeQuarterlyData, stocksNotExisting  



def update_income_statement_annual(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockIncomeAnnuallyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_income_statement_annual(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            stockBalanceData.insert(0, "Symbol", stock)

            allStockIncomeAnnuallyData = pd.concat([allStockIncomeAnnuallyData, stockBalanceData], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        
    return allStockIncomeAnnuallyData, stocksNotExisting  


def updateCashflowStatementAnually(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockCashflowAnuallyData = pd.DataFrame()
    counter = 0
        
    for stock in updateNotExistingSymbols:
        
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            stockBalanceData = FundamentalDataAPIPull.get_cash_flow_annual(stock)[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            stockBalanceData["Downloaded_at_datetime"] = timestampToday()
            stockBalanceData["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(stockBalanceData, pd.DataFrame):
                stockBalanceData = pd.DataFrame(stockBalanceData)
            
            stockBalanceData.insert(0, "Symbol", stock)

            allStockCashflowAnuallyData = pd.concat([allStockCashflowAnuallyData, stockBalanceData], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        
    return allStockCashflowAnuallyData, stocksNotExisting 
    


def getTimeSeriesData(updateNotExistingSymbols):
    stocksNotExisting = []
    allStockTimeSeries = pd.DataFrame()
    counter = 0

    for stock in updateNotExistingSymbols:
        try:
            #the API command get_company_overview retrives stock data like Revenue, Cashflow, Ebit, etc. 
            dailySeriesPerStock = TimeSeriesPull.get_daily(stock, outputsize="full")[0]
            #The function identifyLastColumnWithContents checks the number of the last column with content in it, then two columns with timestamps are appended
            dailySeriesPerStock["Downloaded_at_datetime"] = timestampToday()
            dailySeriesPerStock["Downloaded_at_quarter"] = timestampQuarter()
            
            # Check and if not Dataframe datatype, convert to Dataframe
            if not isinstance(dailySeriesPerStock, pd.DataFrame):
                dailySeriesPerStock = pd.DataFrame(dailySeriesPerStock)
            
            # Convert index to a column to preserve the datetime of the actual stock value
            dailySeriesPerStock.reset_index(inplace=True)

            # Rename the column to Date
            dailySeriesPerStock.rename(columns={'index':'Date'}, inplace=True)
            dailySeriesPerStock["date"] = dailySeriesPerStock["date"].dt.date

            dailySeriesPerStock.insert(0, "Symbol", stock)

            allStockTimeSeries = pd.concat([allStockTimeSeries, dailySeriesPerStock], ignore_index=True)
            APIRequestDelay()
        
        except ValueError:
            stocksNotExisting.append(stock)
            print("Error" + str(counter) + " " + str(stock))
            counter = counter + 1
            pass
        

    return allStockTimeSeries, stocksNotExisting  


"""
def refresh(excel_data, refresh_list):
    today = pd.to_datetime(dt.datetime.today())
    today_quarter = int(today.quarter)
    for elem in refresh_list:
        try:
            stock1 = fd.get_company_overview(elem)[0]
            stock1.insert(59, "Downloaded_at", pd.to_datetime(dt.datetime.today()))
            stock1.insert(60, "Downloaded_quarter", today_quarter)
            excel_data.loc[excel_data["Symbol"] == elem] = stock_transform(stock1)
            time.sleep(time_delay)
        except ValueError:
            excel_data_refreshed = excel_data
            return excel_data_ref
    excel_data_refreshed = excel_data
    return excel_data_ref

def load_stock_price_yf(stock_list_new):
    stock_price_df = yf.download(stock_list_new, period="1d")
    stock_price_df_t = stock_price_df["Close"].T
    stock_price_df_t.insert(1, "Date", pd.to_datetime(dt.datetime.today()))
    stock_price_df_t1 = stock_price_df_t.reset_index()
    stock_price_df_t1.columns = ["Symbol","Price","Date"]
    
    return stock_price_df_t1
"""
           
def writeToExcel(mainExcel, worksheet):
    
    excelPath = "output.xlsx"

    book = load_workbook(excelPath)

    balanceSheet = book[worksheet]

    for row in dataframe_to_rows(mainExcel, index=False, header=False):
        balanceSheet.append(row)

    # Speichere die Änderungen
    book.save(excelPath)


def readFromDataBase(database):
    # Connection to Database
    conn = sql.connect("mainDatabase.db")

    try:
        mainSheetDataframe = pd.read_sql(f"SELECT Symbol, Downloaded_at_quarter FROM {database}", conn)
        
        excelSymbolsExisting = list(set(mainSheetDataframe["Symbol"]))
        excelQuartersExisting = list(mainSheetDataframe["Downloaded_at_quarter"])
        
        return excelSymbolsExisting, excelQuartersExisting

    except:

        excelSymbolsExisting = ["A"]
        excelQuartersExisting = timestampQuarter()

        return excelSymbolsExisting, excelQuartersExisting

    finally:
        conn.close()


def writeToDataBase(mainExcel, database):

    conn = sql.connect("mainDatabase.db")
    mainExcel.to_sql(database, conn, if_exists="append", index=False)
    

    conn.close()


def sortDatabaseBySymbolName(table):
    
    conn = sql.connect("mainDatabase.db")

    cursor = conn.cursor()

    # Create temp table differently sorted
    cursor.execute(f"CREATE TABLE sorted_{table} AS SELECT * FROM {table} ORDER BY Symbol ASC")
    
    # Delete old table
    cursor.execute(f"DROP TABLE {table}")
    
    cursor.execute(f"ALTER TABLE sorted_{table} RENAME TO {table}")


    conn.commit()

    conn.close()



"""

def write_to_excel_daily_stock_price(result):
    book = load_workbook("output.xlsx")
    book_stock_price = book["Daily_Stock_Price"]

    writer = pd.ExcelWriter("output.xlsx", engine='openpyxl') 

    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
  
    result.to_excel(writer, index=False, header=True, sheet_name="Daily_Stock_Price")
    writer.save() 

"""

# NOT USED: Implementation as is, leads to deletion of Symbol entries in Stock_symbol_list, because the API request does get canceled from provide due to exhaustion of free API requests per day.
# TO-DO: When this particular error comes back from server, it is needed to ignore this delete func.
def deleteNoneUpdatableSymbols(symbol_list, stocksNotExisting):
    filename = "Stock_symbols_list.xlsx"
    sheetname = "Overview"
     
    # Compare symbol list against stocks not existing anymore at the stock exchange list
    filtered_symbols = [symbol for symbol in symbol_list if symbol not in stocksNotExisting]

    excelDataFrame = pd.DataFrame(filtered_symbols, columns=[0])

    excelDataFrame.to_excel(filename, sheet_name=sheetname, index=False, header=False)







"""

def calc_price_per_share():
    book = load_workbook("output.xlsx")
    book_overview = book["Overview"]
    book_balance_quarterly = book["Balance_Quarterly"]
    book_income_quarterly = book["Income_Quarterly"]
    book_stock_price = book["Daily_Stock_Price"]

    data_balance = book_balance.values
    data_income = book_income.values
    data_stock_price = book_stock_price.values


    header_bal = next(data_balance)
    header_inc = next(data_income)
    header_pri = next(data_stock_price)

    table_bal = pd.DataFrame(data_balance,columns=header_bal, dtype="int64")
    table_bal["commonStockSharesOutstanding"][table_bal["commonStockSharesOutstanding"]=="None"] = 1
    table_inc = pd.DataFrame(data_income,columns=header_inc, dtype="int64")
    table_inc["netIncome"][table_inc["netIncome"]=="None"] = 1
    table_pri = pd.DataFrame(data_stock_price,columns=header_pri, dtype="int64")

    table_pri = table_pri[["Symbol","Price"]]
    table_pri = table_pri.dropna()
    amount_stocks = table_bal["commonStockSharesOutstanding"]
    net_income = table_inc["netIncome"]

    amount_stocks = prep_df_for_calc(amount_stocks)
    net_income = prep_df_for_calc(net_income)

    multipliers = pd.DataFrame(columns=["Symbol", "fiscalDateEnding","EarningsPerShare","SharePrice"])
    multipliers["EarningsPerShare"] = net_income/amount_stocks*4
    multipliers["Symbol"] = table_inc["Symbol"]
    multipliers["fiscalDateEnding"] = table_inc["fiscalDateEnding"]

    for pos,elem in enumerate(multipliers["Symbol"]):
        for pri_elem in table_pri["Symbol"]:
            if elem == pri_elem:
                multipliers["SharePrice"][pos] = table_pri["Price"][table_pri["Symbol"]==elem].values[0]
    
    return multipliers
"""


   